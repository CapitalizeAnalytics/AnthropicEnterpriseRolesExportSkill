#!/usr/bin/env python3
"""
collect_roles.py — Export Claude Enterprise organization roles to .xlsx files.

Resolves an admin_key (from a config JSON file or environment variable), calls the
Claude Roles API to fetch all custom RBAC roles, and writes detailed .xlsx files for
each role with tabs for:
    1. Membership   - groups assigned to the role, and the members in each
    2. Capabilities - every observed capability, with an Enabled flag
    3. Permissions  - admin permissions with their access level
    4. Connectors   - connectors / tools / scopes with their approval setting
    5. Models       - models available to the role, with a Default flag
    6. Unclassified - any permission row the classifier didn't recognize

Connector permissions carry only opaque IDs (mcpsrv_...) and no endpoint on the
RBAC surface resolves them, so names come from an optional connectors.json mapping;
unnamed IDs are written to a fill-in template in the output directory.

Files are named <Role>_<yyyymmdd_hhmmss>.xlsx, sharing one timestamp per run.

Requires: requests, openpyxl (install with: pip install requests openpyxl)
Python 3 standard library for the rest.

Usage:
    python3 collect_roles.py --config /path/to/config.json [--output RoleDetails]
    ADMIN_API_KEY='...' python3 collect_roles.py [--output RoleDetails]
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterator

try:
    import requests
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit(
        "This script requires 'requests' and 'openpyxl'.\n"
        "Install them with: pip install requests openpyxl"
    )

SKILL_DIR = Path(__file__).resolve().parent.parent
EXAMPLE_PATH = SKILL_DIR / "config.example.json"

BASE = "https://api.anthropic.com/v1/organizations"
BETA_HEADER = "ce-user-management-2026-07-13"

ENV_ADMIN_KEY = "ADMIN_API_KEY"
ENV_API_BASE_URL = "ROLES_API_BASE_URL"

DEFAULT_API_BASE_URL = "https://api.anthropic.com"

# Rate limit is 100 req/min org-wide. Stay under it.
MIN_INTERVAL = 0.65
_last_call = 0.0


# --------------------------------------------------------------------------
# Configuration
# --------------------------------------------------------------------------

def load_example_placeholders() -> dict:
    """The placeholder values live in config.example.json, not duplicated here."""
    try:
        example = json.loads(EXAMPLE_PATH.read_text(encoding="utf-8"))
        if isinstance(example, dict):
            return example
    except (OSError, json.JSONDecodeError):
        pass
    return {
        "admin_key": "sk-ant-admin01-REPLACE_ME",
    }


def load_config_file(path: Path) -> dict:
    if not path.exists():
        sys.exit(
            f"Config file not found at '{path}'.\n"
            f"Either create it (copy config.example.json and fill in admin_key),\n"
            f"or supply the value via the {ENV_ADMIN_KEY} environment variable."
        )
    try:
        config = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        sys.exit(f"'{path}' is not valid JSON: {e}")
    if not isinstance(config, dict):
        sys.exit(f"'{path}' must contain a JSON object at the top level.")
    return config


def resolve_settings_source(args) -> dict:
    """Merge config file and environment into one resolved config.

    Precedence, highest first: environment variables, config file. The override order
    lets someone run this once with values typed in a chat session without having to
    write a credential to disk first.

    Note: there is no --admin-key flag. Command-line arguments are visible to other
    users via the process table and land in shell history; the key is accepted only
    through the environment or the config file for that reason.
    """
    config = load_config_file(Path(args.config)) if args.config else {}

    admin_key = os.environ.get(ENV_ADMIN_KEY) or config.get("admin_key")
    api_base_url = (
        os.environ.get(ENV_API_BASE_URL)
        or config.get("api_base_url")
        or DEFAULT_API_BASE_URL
    )

    resolved = {
        "admin_key": (admin_key or "").strip(),
        "api_base_url": (api_base_url or "").strip(),
    }

    # Only admin_key carries a placeholder. api_base_url in config.example.json is a
    # real working default, so comparing it against the example would reject every
    # config that sensibly left it alone.
    example = load_example_placeholders()
    if not resolved["admin_key"] or resolved["admin_key"] == str(example.get("admin_key", "")):
        sys.exit(
            "Missing or placeholder value for 'admin_key'. "
            f"Set it in config.json, via the {ENV_ADMIN_KEY} environment variable, "
            "or both."
        )
    if not resolved["api_base_url"]:
        sys.exit(
            "Missing value for 'api_base_url'. Leave it out entirely to use the default "
            f"({DEFAULT_API_BASE_URL}), or set a real endpoint."
        )

    return resolved


def save_config(config: dict, path: Path) -> None:
    """Write resolved config to disk (for reuse on next invocation)."""
    payload = {
        "admin_key": config.get("admin_key"),
        "api_base_url": config.get("api_base_url"),
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


# --------------------------------------------------------------------------
# Connector names
# --------------------------------------------------------------------------

CONNECTOR_MAP_NAME = "connectors.json"
CONNECTOR_EXAMPLE_NAME = "connectors.example.json"
UNRESOLVED_STUB_NAME = "connectors.unresolved.json"
# Keys carrying this marker come from connectors.example.json copied but never edited.
CONNECTOR_PLACEHOLDER = "EXAMPLE_REPLACE_ME"


def load_connector_names(explicit: str | None) -> dict[str, str]:
    """Map connector IDs (mcpsrv_...) to human-readable names.

    No endpoint on the RBAC surface lists an organization's connectors, so the
    permission rows carry bare IDs with nothing to resolve them against. The
    analytics API does expose /v1/organizations/analytics/connectors (needs the
    read:analytics scope), but it reports *usage*, not a registry: names there are
    normalized across sources ("mcp-atlassian" and "Atlassian MCP server" both
    become "atlassian"), and whether rows carry the mcpsrv_ ID needed to join back
    to these permissions is unconfirmed — the endpoint returned no rows on any date
    tested. So this reads an operator-maintained mapping instead. Accepts either a
    flat {id: name} object or {"connectors": {id: name}}.
    """
    candidates = [Path(explicit)] if explicit else [
        Path.cwd() / CONNECTOR_MAP_NAME,
        SKILL_DIR / CONNECTOR_MAP_NAME,
    ]
    found_any = False
    for path in candidates:
        if not path.exists():
            continue
        found_any = True
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as e:
            sys.stderr.write(f"  WARNING: ignoring '{path}' ({e})\n")
            continue
        if not isinstance(data, dict):
            sys.stderr.write(f"  WARNING: ignoring '{path}' (expected a JSON object)\n")
            continue
        inner = data.get("connectors") if isinstance(data.get("connectors"), dict) else data
        mapping = {
            str(k): str(v) for k, v in inner.items()
            if v and not str(k).startswith("_") and CONNECTOR_PLACEHOLDER not in str(k)
        }
        if mapping:
            print(f"  connector names: {len(mapping)} from {path}")
            return mapping
        # A file that resolved nothing (still all placeholders, or emptied out) must not
        # shadow a real mapping further down the search path.
        sys.stderr.write(f"  note: '{path}' supplied no connector names; looking further\n")
    if explicit and not found_any:
        sys.exit(f"Connector map not found at '{explicit}'.")
    return {}


def write_unresolved_stub(ids: set[str], names: dict[str, str], out_dir: Path) -> Path | None:
    """Drop a fill-in-the-blanks template for IDs we couldn't name."""
    missing = sorted(i for i in ids if i not in names)
    if not missing:
        return None
    payload = {
        "_comment": (
            "Rename this file to connectors.json (in the working directory or the "
            "skill directory) and replace each empty value with the connector's "
            "display name from claude.ai > Organization settings > Connectors. "
            "Existing names already resolved are included for reference."
        ),
        "connectors": {**{i: names[i] for i in sorted(names)}, **{i: "" for i in missing}},
    }
    path = out_dir / UNRESOLVED_STUB_NAME
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return path


# --------------------------------------------------------------------------
# HTTP
# --------------------------------------------------------------------------

class Client:
    def __init__(self, api_key: str, api_base: str = DEFAULT_API_BASE_URL):
        self.api_base = api_base
        self.s = requests.Session()
        self.s.headers.update({
            "x-api-key": api_key,
            "anthropic-beta": BETA_HEADER,
        })

    def get(self, path: str, **params) -> dict:
        """GET with throttling and 429/5xx backoff."""
        global _last_call
        url = f"{self.api_base}/v1/organizations{path}"
        for attempt in range(6):
            gap = time.monotonic() - _last_call
            if gap < MIN_INTERVAL:
                time.sleep(MIN_INTERVAL - gap)
            _last_call = time.monotonic()

            r = self.s.get(url, params={k: v for k, v in params.items() if v is not None})
            if r.status_code == 429 or r.status_code >= 500:
                wait = float(r.headers.get("retry-after", 2 ** attempt))
                sys.stderr.write(f"  {r.status_code} on {path}, retrying in {wait:.0f}s\n")
                time.sleep(wait)
                continue
            if not r.ok:
                raise RuntimeError(f"{r.status_code} {path}: {r.text[:400]}")
            return r.json()
        raise RuntimeError(f"gave up after retries: {path}")

    def paginate(self, path: str, limit: int = 100) -> Iterator[dict]:
        """Cursor pagination: feed next_page back in as `page` until it's null."""
        page = None
        while True:
            body = self.get(path, limit=limit, page=page)
            yield from body.get("data", [])
            page = body.get("next_page")
            if not page:
                return


# --------------------------------------------------------------------------
# Classification
# --------------------------------------------------------------------------

CONNECTOR_RESOURCES = {"connector", "connector_tool", "connector_scope", "all_connectors"}
BLANKET_ACTIONS = {"capability_access_all", "capability_access_all_ga"}
ACCESS_RANK = {"no_access": 0, "view": 1, "can_view": 1, "manage": 2, "can_manage": 2}

# How a connector grant is described once its rows are merged. Ordered most- to
# least-permissive; the first action present wins, so a connector carrying both
# "use" and "always_allow" reports as auto-approved rather than merely allowed.
CONNECTOR_ACTION_LABELS = {
    "always_allow": "Always allow (no prompt)",
    "interactive": "Ask each time",
    "use": "Allowed",
    "no_access": "Blocked",
}
CONNECTOR_ACTION_ORDER = ("no_access", "always_allow", "interactive", "use")


def bucket(perm: dict) -> str:
    res = perm.get("resource") or {}
    rtype = res.get("type", "")
    action = perm.get("action", "")

    if rtype in CONNECTOR_RESOURCES:
        return "connectors"
    if "model" in rtype or action.startswith("model"):
        return "models"
    if rtype == "organization":
        if action.startswith("permission_"):
            return "permissions"
        # Everything else scoped to the organization is a capability. The API emits
        # these both prefixed (capability_access_all) and bare (chat, skill_creation),
        # so matching on a "capability_" prefix would strand the bare ones.
        return "capabilities"
    return "unclassified"


def prettify(token: str) -> str:
    """capability_web_search -> Web Search"""
    for prefix in ("capability_", "permission_", "model_"):
        if token.startswith(prefix):
            token = token[len(prefix):]
            break
    return re.sub(r"[_\-]+", " ", token).strip().title()


# --------------------------------------------------------------------------
# Fetch
# --------------------------------------------------------------------------

def fetch_roles(c: Client) -> list[dict]:
    return list(c.paginate("/rbac_roles"))


def fetch_groups_by_role(c: Client) -> dict[str, list[dict]]:
    """Reverse-index groups onto the roles attached to them, members included."""
    index: dict[str, list[dict]] = defaultdict(list)
    for g in c.paginate("/rbac_groups"):
        roles = g.get("roles")
        if roles is None:
            # Degraded read, not an empty group. Retry once before trusting it.
            retry = c.get(f"/rbac_groups/{g['id']}")
            roles = retry.get("roles")
            if roles is None:
                sys.stderr.write(
                    f"  WARNING: group '{g.get('name')}' returned null roles twice; "
                    f"its role assignments are missing from this export\n"
                )
                continue
        if not roles:
            continue
        members = list(c.paginate(f"/rbac_groups/{g['id']}/members"))
        for role_id in roles:
            index[role_id].append({**g, "_members": members})
    return index


def fetch_permissions(c: Client, role_id: str) -> list[dict]:
    return list(c.paginate(f"/rbac_roles/{role_id}/permissions"))


# --------------------------------------------------------------------------
# Workbook
# --------------------------------------------------------------------------

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
BODY = Font(name=FONT, size=10)
GROUP_FONT = Font(name=FONT, bold=True, size=10)


def write_sheet(ws, headers: list[str], rows: list[list[Any]], bold_rows: set[int] = frozenset()):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for i, row in enumerate(rows):
        ws.append(row)
        font = GROUP_FONT if i in bold_rows else BODY
        for cell in ws[ws.max_row]:
            cell.font = font

    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        widest = max(
            [len(str(headers[col - 1]))] +
            [len(str(r[col - 1])) for r in rows if col - 1 < len(r) and r[col - 1] is not None]
        )
        ws.column_dimensions[letter].width = min(max(widest + 3, 12), 60)


def safe_name(name: str) -> str:
    cleaned = re.sub(r'[\\/*?:\[\]]', "-", name).strip() or "role"
    return cleaned[:100]


def build_workbook(role: dict, groups: list[dict], perms: list[dict],
                   universe: dict[str, set[str]], out_dir: Path,
                   connector_names: dict[str, str] | None = None,
                   stamp: str = "") -> Path:
    connector_names = connector_names or {}
    wb = Workbook()
    wb.remove(wb.active)

    buckets: dict[str, list[dict]] = defaultdict(list)
    for p in perms:
        buckets[bucket(p)].append(p)

    # --- 1. Membership -----------------------------------------------------
    rows, bold = [], set()
    for g in sorted(groups, key=lambda x: x.get("name", "")):
        members = g.get("_members", [])
        source = "Identity provider (SCIM)" if g.get("source_type") == "scim" else "Created in Claude"
        rows.append([g.get("name"), "", "", source, len(members)])
        bold.add(len(rows) - 1)
        for m in sorted(members, key=lambda x: (x.get("email") or "")):
            rows.append(["", m.get("email"), m.get("user_id"), "", ""])
    if not rows:
        rows = [["(no groups assigned to this role)", "", "", "", ""]]
    write_sheet(wb.create_sheet("Membership"),
                ["Group", "Member Email", "User ID", "Group Source", "Member Count"],
                rows, bold)

    # --- 2. Capabilities ---------------------------------------------------
    granted = {p["action"] for p in buckets["capabilities"]}
    blanket = granted & BLANKET_ACTIONS
    rows = []
    if blanket:
        label = "All capabilities" if "capability_access_all" in blanket else "All generally available capabilities"
        rows.append([label, "BLANKET GRANT", "Yes",
                     "Covers every capability in scope. Excludes model access and admin permissions."])
    for action in sorted(universe["capabilities"] - BLANKET_ACTIONS):
        enabled = "Yes" if (action in granted or blanket) else "No"
        note = "Granted via blanket grant" if (blanket and action not in granted) else ""
        rows.append([prettify(action), action, enabled, note])
    write_sheet(wb.create_sheet("Capabilities"),
                ["Capability", "Action Key", "Enabled", "Notes"], rows)

    # --- 3. Permissions ----------------------------------------------------
    by_action: dict[str, str] = {}
    for p in buckets["permissions"]:
        a, lvl = p["action"], str(p.get("access") or p.get("level") or "granted")
        if ACCESS_RANK.get(lvl, 1) >= ACCESS_RANK.get(by_action.get(a, ""), -1):
            by_action[a] = lvl
    rows = [[prettify(a), a, prettify(by_action.get(a, "no_access"))]
            for a in sorted(universe["permissions"])]
    write_sheet(wb.create_sheet("Permissions"),
                ["Admin Permission", "Action Key", "Access Type"], rows)

    # --- 4. Connectors -----------------------------------------------------
    # One row per distinct target. The API emits access and approval as separate
    # permission rows for the same connector ("use" alongside "always_allow"), which
    # read as duplicates in a spreadsheet, so collapse them and derive one setting.
    merged: dict[tuple, set[str]] = defaultdict(set)
    for p in buckets["connectors"]:
        res = p.get("resource") or {}
        key = (res.get("type", ""), res.get("connector_id") or "",
               res.get("tool_name") or "", res.get("scope") or "")
        merged[key].add(p.get("action", ""))

    def sort_key(item):
        (rtype, cid, _tool, _scope), _actions = item
        name = connector_names.get(cid, "")
        # Org-wide default first, then named connectors A-Z, then unnamed IDs.
        tier = 0 if rtype == "all_connectors" else (1 if name else 2)
        return (tier, name.lower(), cid)

    rows = []
    for (rtype, cid, tool, scope_name), actions in sorted(merged.items(), key=sort_key):
        scope = {
            "all_connectors": "All connectors",
            "connector": "Whole connector",
            "connector_tool": "Individual tool",
            "connector_scope": "OAuth scope",
        }.get(rtype, rtype)
        # Only name a tool or scope here. Falling back to the connector ID repeats
        # the ID column and tells the reader nothing new.
        if tool:
            target = tool
        elif scope_name:
            target = scope_name
        elif rtype == "all_connectors":
            target = "(every connector)"
        else:
            target = "(entire connector)"
        if rtype == "all_connectors":
            name = "All connectors (organization-wide default)"
        else:
            name = connector_names.get(cid) or "(unnamed - add to connectors.json)"
        approval = next((CONNECTOR_ACTION_LABELS[a] for a in CONNECTOR_ACTION_ORDER
                         if a in actions), None)
        if approval is None:
            approval = ", ".join(prettify(a) for a in sorted(actions)) or "-"
        rows.append([name, cid or "-", scope, target, approval,
                     ", ".join(sorted(actions))])
    if not rows:
        rows = [["(no connector grants - all connectors blocked for this role)",
                 "", "", "", "", ""]]
    write_sheet(wb.create_sheet("Connectors"),
                ["Connector", "Connector ID", "Scope", "Tool / Scope",
                 "Approval Setting", "Granted Actions"], rows)

    # --- 5. Models ---------------------------------------------------------
    rows = []
    for p in buckets["models"]:
        res = p.get("resource") or {}
        model = res.get("model") or res.get("model_id") or res.get("name") or p.get("action", "")
        rows.append([
            model,
            "Yes",
            "Yes" if (res.get("is_default") or p.get("default")) else "",
            res.get("max_effort") or p.get("max_effort") or "",
        ])
    if not rows:
        rows = [["(no model rows returned - see README on reading model access)", "", "", ""]]
    write_sheet(wb.create_sheet("Models"),
                ["Model", "Available", "Default", "Max Effort Level"], rows)

    # --- 6. Unclassified ---------------------------------------------------
    if buckets["unclassified"]:
        rows = [[(p.get("resource") or {}).get("type", ""), p.get("action", ""),
                 json.dumps(p.get("resource"), sort_keys=True)]
                for p in buckets["unclassified"]]
        write_sheet(wb.create_sheet("Unclassified"),
                    ["Resource Type", "Action", "Raw Resource"], rows)

    base = safe_name(role.get("name", role["id"]))
    path = out_dir / (f"{base}_{stamp}.xlsx" if stamp else f"{base}.xlsx")
    wb.save(path)
    return path


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser(
        description="Export Claude Enterprise organization roles to .xlsx files."
    )
    ap.add_argument(
        "--config",
        default=None,
        help="Path to config.json (default: ./config.json, then <skill-dir>/config.json)"
    )
    ap.add_argument(
        "--output",
        default="RoleDetails",
        help="Output directory for .xlsx files (default: RoleDetails)"
    )
    ap.add_argument(
        "--save-config",
        default=None,
        help="Save resolved config to this path for reuse (only via environment/CLI, not from existing config)"
    )
    ap.add_argument(
        "--no-groups",
        action="store_true",
        help="Skip the Membership tab (use if your key isn't scoped for all linked orgs)"
    )
    ap.add_argument(
        "--connectors",
        default=None,
        help=f"Path to a {{connector_id: name}} map (default: ./{CONNECTOR_MAP_NAME}, "
             f"then <skill-dir>/{CONNECTOR_MAP_NAME})"
    )
    args = ap.parse_args()

    # Resolve configuration
    config = resolve_settings_source(args)
    api_base = config.get("api_base_url") or DEFAULT_API_BASE_URL

    out = Path(args.output)
    out.mkdir(parents=True, exist_ok=True)
    c = Client(config["admin_key"], api_base)

    try:
        print("Fetching roles...")
        roles = fetch_roles(c)
        print(f"  {len(roles)} role(s)")
        if not roles:
            print("No custom roles found. Confirm the org is on Enterprise and has custom roles.")
            return 0

        groups_by_role: dict[str, list[dict]] = {}
        if not args.no_groups:
            print("Fetching groups and members...")
            try:
                groups_by_role = fetch_groups_by_role(c)
            except RuntimeError as e:
                sys.stderr.write(f"  Group fetch failed ({e}).\n"
                                 f"  Continuing without Membership data. "
                                 f"Groups need read:rbac_groups on an all-linked-orgs key.\n")

        print("Fetching permissions per role...")
        perms_by_role = {}
        for r in roles:
            perms_by_role[r["id"]] = fetch_permissions(c, r["id"])
            print(f"  {r.get('name')}: {len(perms_by_role[r['id']])} permission row(s)")

        # Observed universe across every role
        universe = {"capabilities": set(), "permissions": set()}
        for perms in perms_by_role.values():
            for p in perms:
                b = bucket(p)
                if b in universe:
                    universe[b].add(p["action"])

        connector_names = load_connector_names(args.connectors)
        connector_ids = {
            (p.get("resource") or {}).get("connector_id")
            for perms in perms_by_role.values() for p in perms
            if bucket(p) == "connectors" and (p.get("resource") or {}).get("connector_id")
        }

        # One stamp for the whole run, so a single export's files sort together.
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        print("Writing workbooks...")
        for r in roles:
            p = build_workbook(r, groups_by_role.get(r["id"], []),
                               perms_by_role[r["id"]], universe, out,
                               connector_names, stamp)
            print(f"  {p}")

        stub = write_unresolved_stub(connector_ids, connector_names, out)
        if stub:
            unnamed = len(connector_ids - set(connector_names))
            print(f"\n{unnamed} connector ID(s) had no name available. The Admin API has no "
                  f"endpoint that lists connectors,\nso names must be supplied manually. "
                  f"A template is at:\n  {stub}\nFill in the names, save it as "
                  f"'{CONNECTOR_MAP_NAME}', and re-run to label the Connectors tab.")

        # Save config if requested
        if args.save_config:
            config_path = Path(args.save_config)
            save_config(config, config_path)
            print(f"\nConfig saved to {config_path}")

        print(f"\nDone. {len(roles)} role(s) exported to {out}")
        return 0

    except RuntimeError as e:
        sys.stderr.write(f"Error: {e}\n")
        return 1


if __name__ == "__main__":
    sys.exit(main())
